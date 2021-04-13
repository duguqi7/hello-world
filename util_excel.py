#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import openpyxl
import param
import sys
import os

unified_wb_name = '固件识别码统计信息'

def get_vernum(version):
	if False == isinstance(version, str):
		print(sys._getframe().f_lineno, 'version error', version)
		return ''
	elif version[0] != 'V' and version[0] != 'v':
		print(sys._getframe().f_lineno, 'version type error', version)
		return ''
	start_postion = 1;
	end_position = version.find('.', start_postion)
	if -1 == end_position:
		print(sys._getframe().f_lineno, 'version error', version)
		return ''
	main = version[start_postion:end_position].rjust(2, '0')
	start_postion = 1+end_position;
	end_position = version.find('.', start_postion)
	if -1 == end_position:
		print(sys._getframe().f_lineno, 'version error', version)
		return ''
	sub = version[start_postion:end_position].rjust(2, '0')
	start_postion = 1+end_position;
	end_position = version.find('build')
	if -1 == end_position:
		aux = version[start_postion:].rjust(4, '0')
	else:
		aux = version[start_postion:end_position].rjust(4, '0')
	return (main + sub + aux)

def get_date(version):
	if False == isinstance(version, str):
		print('version error', version)
		return ''
	start_postion = version.find('build') + len('build')
	year = int(version[start_postion:start_postion+2])
	month = int(version[start_postion+2:start_postion+4])
	day = int(version[start_postion+4:])
	if year < 20 or month > 12 or day > 31:
		print('version simple check error', version)
		return ''
	#组建版本日期
	date = hex(year)[2:4].rjust(2, '0') + hex(month)[2:4].rjust(2, '0') + hex(day)[2:4].rjust(2, '0')
	return date.rjust(8, '0')


def update(version, date, filename):
	if False == isinstance(version, str) or \
	   '' == version or \
	   False == isinstance(date, str) or \
	   '' == date or \
	   False == isinstance(filename, str) or \
	   '' == filename:
		print('222 line:', sys._getframe().f_lineno, 'param error', version, date, filename)
		return None
	wb = openpyxl.load_workbook(filename)
	if unified_wb_name not in wb.sheetnames:
		print('can not find sheet')
		return None
	else:
		ws = wb[unified_wb_name]

	for row in ws.values:
		temp_row = row
	modify_row = list(temp_row)
	modify_row[6] = version[1:]
	modify_row[7] = int(date[5:])
	vernum = get_vernum(version)
	datenum = get_date(date)
	firmcode = modify_row[11][:64] + vernum + datenum + modify_row[11][80:]
	modify_row[11] = firmcode
	ws.append(modify_row)
	datavalidations = ws.data_validations.dataValidation
	for validation in datavalidations:
		validation.ranges.ranges[0].max_row += 1
	wb.save(filename)
	split_name, split_type = os.path.splitext(filename)
	verposition = split_name.find('V')
	dateposition = split_name.find('build')
	newname = split_name[:verposition] + version + '_' + date + '_dst_dev_codes' + split_type
	os.rename(filename, newname)
	print('update %s ok' % newname)
	return 'OK'